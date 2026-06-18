from __future__ import annotations

import json
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
DATA_WORKBOOK_CANDIDATES = [
    ROOT / "Data.xlsx",
    ROOT / "data.xlsx",
    ROOT / "Data.xlsm",
    ROOT / "data.xlsm",
]
OUTPUT = Path(__file__).resolve().parents[1] / "data" / "mill_weight.json"
SHEET_NAME = "SPC"

FIELDS = [
    "wpDocNo",
    "wpDocDate",
    "wpctCode",
    "ctinit",
    "ctfname",
    "ctlname",
    "wpCarLicense",
    "wpNetWeight",
    "wpGradeNew",
    "wpproduct",
    "wppriceperunit",
    "wptotalpay",
    "wpRspo",
]


def first_existing(paths: list[Path]) -> Path:
    for path in paths:
        if path.is_file():
            return path
    choices = ", ".join(str(path) for path in paths)
    raise FileNotFoundError(f"Missing Data workbook. Tried: {choices}")


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, str):
        text = value.strip()
        return text if text else None
    return value


def number(value: Any) -> float:
    try:
        if value in (None, ""):
            return 0.0
        return float(str(value).replace(",", ""))
    except (TypeError, ValueError):
        return 0.0


def iso_day(value: Any) -> str | None:
    cleaned = clean_value(value)
    if not cleaned:
        return None
    return str(cleaned)[:10]


def display_datetime(value: Any) -> str | None:
    if isinstance(value, datetime):
        return f"{value:%d/%m/%Y} {value.hour}:{value:%M}"
    cleaned = clean_value(value)
    return str(cleaned) if cleaned else None


def customer_name(row: dict[str, Any]) -> str:
    return " ".join(str(row.get(field) or "").strip() for field in ["ctinit", "ctfname", "ctlname"] if row.get(field)).strip()


def read_records(workbook: Path) -> tuple[list[dict[str, Any]], list[str]]:
    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise KeyError(f"Missing sheet {SHEET_NAME} in {workbook.name}")
    ws = wb[SHEET_NAME]
    headers = [clean_value(value) for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {str(header): pos for pos, header in enumerate(headers) if header}
    missing = [field for field in FIELDS if field not in index]
    rows: list[dict[str, Any]] = []

    for source_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in values):
            continue
        raw_doc_date = values[index["wpDocDate"]] if "wpDocDate" in index and index["wpDocDate"] < len(values) else None
        row: dict[str, Any] = {"sourceRow": source_row}
        for field in FIELDS:
            raw = values[index[field]] if field in index and index[field] < len(values) else None
            value = clean_value(raw)
            if field in {"wpNetWeight", "wppriceperunit", "wptotalpay"}:
                value = number(raw)
            row[field] = value
        row["wpDocDateText"] = display_datetime(raw_doc_date)
        row["date"] = iso_day(raw_doc_date)
        row["docKey"] = str(row.get("wpDocNo") or "").strip()
        row["customerName"] = customer_name(row)
        rows.append(row)

    return rows, missing


def main() -> None:
    workbook = first_existing(DATA_WORKBOOK_CANDIDATES)
    records, missing = read_records(workbook)
    dates = [row["date"] for row in records if row.get("date")]
    payload = {
        "source": {
            "workbook": workbook.name,
            "sheet": SHEET_NAME,
            "rowCount": len(records),
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "dateMin": min(dates) if dates else None,
            "dateMax": max(dates) if dates else None,
            "description": "ข้อมูลน้ำหนักปลายทางโรงงาน SPC จากชีต SPC ใน Data.xlsx",
            "missingFields": missing,
        },
        "records": records,
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(records):,} SPC records to {OUTPUT}")


if __name__ == "__main__":
    main()
