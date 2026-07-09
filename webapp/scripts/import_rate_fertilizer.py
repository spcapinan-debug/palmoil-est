from __future__ import annotations

import argparse
import json
import re
import urllib.error
import urllib.parse
import urllib.request
from calendar import monthrange
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_WORKBOOK = ROOT / "Master Data" / "Rate Fertilizer.xlsx"
DEFAULT_API = "https://palmoil-est.vercel.app/api/farm-tables"
SOURCE_FILE = "Rate Fertilizer.xlsx"
SOURCE_SHEET = "Fertilizer"
FISCAL_YEAR = "2569"
ACTIVITY_CODE = "MN02"
ACTIVITY_NAME = "\u0e43\u0e2a\u0e48\u0e1b\u0e38\u0e4b\u0e22"

THAI_MONTHS = {
    "\u0e21.\u0e04.": 1,
    "\u0e01.\u0e1e.": 2,
    "\u0e21\u0e35.\u0e04.": 3,
    "\u0e40\u0e21.\u0e22.": 4,
    "\u0e1e.\u0e04.": 5,
    "\u0e21\u0e34.\u0e22.": 6,
    "\u0e01.\u0e04.": 7,
    "\u0e2a.\u0e04.": 8,
    "\u0e01.\u0e22.": 9,
    "\u0e15.\u0e04.": 10,
    "\u0e1e.\u0e22.": 11,
    "\u0e18.\u0e04.": 12,
}

MATERIAL_ALIASES = [
    ("Dolomite", "F-CM-0001"),
    ("\u0e42\u0e14\u0e42\u0e25", "F-CM-0001"),
    ("0-3-0", "F-CM-0004"),
    ("0-0-60", "F-CM-0007"),
    ("46-0-0", "F-CM-0006"),
    ("Borax", "F-CM-0005"),
    ("\u0e1a\u0e2d\u0e41\u0e23\u0e01", "F-CM-0005"),
]


@dataclass
class RateColumn:
    col: int
    month_label: str
    month_no: int
    material_label: str
    material_code: str
    grams_per_tree: float


def norm_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip())


def safe_code(value: Any) -> str:
    text = norm_text(value).upper()
    text = re.sub(r"[^A-Z0-9]+", "-", text).strip("-")
    return text or "X"


def api_request(
    api_base: str,
    method: str = "GET",
    payload: dict[str, Any] | None = None,
    query: dict[str, Any] | None = None,
) -> dict[str, Any]:
    url = api_base
    if query:
        url = f"{url}?{urllib.parse.urlencode(query)}"
    data = None
    headers = {"Content-Type": "application/json"}
    if payload is not None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    request = urllib.request.Request(url, data=data, method=method, headers=headers)
    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as error:
        body = error.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"{method} {url} failed {error.code}: {body}") from error


def load_tables(api_base: str) -> dict[str, list[dict[str, Any]]]:
    table_names = [
        "blocks",
        "activities",
        "activity_groups",
        "materials",
        "budget_years",
        "budget_activity_rates",
        "budget_rate_blocks",
        "budget_rate_materials",
    ]
    payload = api_request(api_base, query={"tables": ",".join(table_names)})
    tables = payload.get("tables") or {}
    return {name: tables.get(name) or [] for name in table_names}


def material_code_for(label: str) -> str:
    lowered = label.lower()
    for needle, code in MATERIAL_ALIASES:
        if needle.lower() in lowered:
            return code
    return ""


def month_range(month_no: int) -> tuple[str, str]:
    start = date(2026, month_no, 1)
    end = date(2026, month_no, monthrange(2026, month_no)[1])
    return start.isoformat(), end.isoformat()


def parse_rate_columns(ws) -> list[RateColumn]:
    columns: list[RateColumn] = []
    current_month = ""
    current_material = ""
    for col in range(2, ws.max_column + 1):
        month = norm_text(ws.cell(4, col).value) or current_month
        material = norm_text(ws.cell(5, col).value) or current_material
        if month:
            current_month = month
        if material:
            current_material = material
        grams = ws.cell(6, col).value
        if not month or not material or not isinstance(grams, (int, float)):
            continue
        month_no = THAI_MONTHS.get(month)
        material_code = material_code_for(material)
        if not month_no or not material_code:
            continue
        columns.append(RateColumn(
            col=col,
            month_label=month,
            month_no=month_no,
            material_label=material,
            material_code=material_code,
            grams_per_tree=float(grams),
        ))
    return columns


def parse_block_quantities(ws, columns: list[RateColumn]) -> dict[int, list[dict[str, Any]]]:
    by_col = {col.col: [] for col in columns}
    current_zone = ""
    for row in range(8, ws.max_row + 1):
        block_code = norm_text(ws.cell(row, 1).value)
        if not block_code:
            continue
        if block_code in {"Lower", "Upper", "Grand Total"}:
            current_zone = block_code if block_code != "Grand Total" else current_zone
            continue
        if not re.search(r"\d", block_code):
            continue
        for col in columns:
            bags = ws.cell(row, col.col).value
            if not isinstance(bags, (int, float)) or bags <= 0:
                continue
            by_col[col.col].append({
                "block_code": block_code,
                "zone_name": current_zone,
                "planned_bags": float(bags),
            })
    return by_col


def find_one(rows: list[dict[str, Any]], *predicates) -> dict[str, Any]:
    for row in rows:
        if all(predicate(row) for predicate in predicates):
            return row
    return {}


def build_block_lookup(blocks: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}
    for block in blocks:
        for key in [block.get("block_code"), block.get("terrain_code"), block.get("block_name"), block.get("area_code")]:
            if key:
                lookup[norm_text(key).upper()] = block
    return lookup


def build_import_rows(
    workbook: Path,
    tables: dict[str, list[dict[str, Any]]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    wb = load_workbook(workbook, data_only=True)
    ws = wb[SOURCE_SHEET]
    rate_columns = parse_rate_columns(ws)
    block_quantities = parse_block_quantities(ws, rate_columns)
    activity = find_one(tables["activities"], lambda row: row.get("activity_code") == ACTIVITY_CODE)
    if not activity:
        raise RuntimeError(f"Activity {ACTIVITY_CODE} not found")
    group = find_one(tables["activity_groups"], lambda row: row.get("id") == activity.get("activity_group_id"))
    year = find_one(tables["budget_years"], lambda row: str(row.get("fiscal_year")) == FISCAL_YEAR)
    if not year:
        raise RuntimeError(f"Budget year {FISCAL_YEAR} not found")

    block_lookup = build_block_lookup(tables["blocks"])
    material_lookup = {row.get("material_code"): row for row in tables["materials"]}
    rates: list[dict[str, Any]] = []
    rate_blocks: list[dict[str, Any]] = []
    rate_materials: list[dict[str, Any]] = []
    warnings: list[str] = []

    seq = 1
    for col in rate_columns:
        blocks_for_rate = block_quantities.get(col.col, [])
        if not blocks_for_rate:
            continue
        material = material_lookup.get(col.material_code) or {}
        start_date, end_date = month_range(col.month_no)
        rate_code = f"R69-MN02-F{seq:03d}"
        rate_id = f"budget-rate-{rate_code.lower()}"
        material_name = material.get("material_name") or col.material_label
        matched_blocks: list[dict[str, Any]] = []
        total_bags = 0.0
        total_trees = 0.0
        total_rai = 0.0
        for item in blocks_for_rate:
            block = block_lookup.get(item["block_code"].upper())
            if not block:
                warnings.append(f"ไม่พบ Block {item['block_code']} ในฐานข้อมูล")
                continue
            matched_blocks.append(block)
            total_bags += item["planned_bags"]
            total_trees += float(block.get("tree_count") or 0)
            total_rai += float(block.get("area_rai") or 0)
            rate_blocks.append({
                "id": f"budget-rate-block-{rate_code.lower()}-{safe_code(item['block_code']).lower()}",
                "budget_rate_id": rate_id,
                "block_id": block.get("id") or "",
                "terrain_code": block.get("block_code") or item["block_code"],
                "block_name": block.get("block_name") or item["block_code"],
                "estate_name": block.get("estate_name") or "",
                "zone_name": block.get("zone_name") or item.get("zone_name") or "",
                "plot_group_code": block.get("plot_group_code") or block.get("plot_group_name") or "",
                "ap_code": block.get("ap_code") or block.get("AP_code") or "",
                "rspo_status": block.get("rspo_status") or "",
                "area_rai": block.get("area_rai") or "",
                "tree_count": block.get("tree_count") or "",
                "status": "active",
                "note": f"{SOURCE_FILE} col {col.col} | แผน {item['planned_bags']:g} กระสอบ",
            })
        if not matched_blocks:
            seq += 1
            continue
        rate_text = f"{material_name} {col.grams_per_tree:g} กรัม/ต้น | {col.month_label} | {total_bags:g} กระสอบ"
        rates.append({
            "id": rate_id,
            "budget_year_id": year.get("id"),
            "fiscal_year": FISCAL_YEAR,
            "rate_code": rate_code,
            "activity_group_name": group.get("group_name") or "การใส่ปุ๋ย",
            "activity_id": activity.get("id"),
            "activity_code": ACTIVITY_CODE,
            "activity_name": activity.get("activity_name") or ACTIVITY_NAME,
            "rate_type": "material",
            "calculation_method": "per_tree",
            "comparison_basis": "tree_count",
            "unit_name": "กรัม/ต้น",
            "rate_amount": 0,
            "rate_text": rate_text,
            "area_scope_type": "block",
            "estate_name": "SPC Kirirat Estate",
            "zone_name": f"{len(matched_blocks)} Block",
            "plot_group_code": "",
            "block_id": "",
            "terrain_code": ", ".join((row.get("block_code") or row.get("terrain_code") or "").strip() for row in matched_blocks[:5]),
            "ap_code": ", ".join(sorted({norm_text(row.get("ap_code") or row.get("AP_code")) for row in matched_blocks if row.get("ap_code") or row.get("AP_code")})[:5]),
            "rspo_status": ", ".join(sorted({norm_text(row.get("rspo_status")) for row in matched_blocks if row.get("rspo_status")})),
            "area_rai": total_rai,
            "tree_count": total_trees,
            "effective_from": start_date,
            "effective_to": end_date,
            "approval_status": "approved",
            "version_no": 1,
            "is_current": "true",
            "source_file": SOURCE_FILE,
            "source_sheet": SOURCE_SHEET,
            "source_column": str(col.col),
            "source_row": 6,
            "mapping_rule": f"{col.month_label}/{col.material_label}/{col.grams_per_tree:g}g/tree",
            "status": "active",
            "note": f"นำเข้าจาก {SOURCE_FILE}: {rate_text}. ใช้วางแผน MN02 ตามรอบปุ๋ยต่อเนื่อง",
        })
        rate_materials.append({
            "id": f"budget-rate-material-{rate_code.lower()}-{safe_code(col.material_code).lower()}",
            "budget_rate_id": rate_id,
            "material_id": material.get("id") or "",
            "material_name": material_name,
            "usage_quantity": col.grams_per_tree,
            "usage_unit": "กรัม",
            "usage_basis": "tree_count",
            "unit_cost": "",
            "amount_per_basis": "",
            "status": "active",
            "note": f"{col.month_label} | {total_bags:g} กระสอบ | {len(matched_blocks)} Block",
        })
        seq += 1
    return rates, rate_blocks, rate_materials, warnings


def post_rows(api_base: str, table: str, rows: list[dict[str, Any]], dry_run: bool) -> int:
    if dry_run or not rows:
        return len(rows)
    payload = api_request(api_base, "POST", {"table": table, "rows": rows, "reason": f"import {SOURCE_FILE}"})
    if not payload.get("ok"):
        raise RuntimeError(f"Import {table} failed: {payload}")
    return int(payload.get("count") or len(payload.get("rows") or rows))


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--api", default=DEFAULT_API)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    workbook = Path(args.workbook)
    if not workbook.exists():
        raise FileNotFoundError(workbook)
    tables = load_tables(args.api)
    rates, rate_blocks, rate_materials, warnings = build_import_rows(workbook, tables)
    counts = {
        "budget_activity_rates": post_rows(args.api, "budget_activity_rates", rates, args.dry_run),
        "budget_rate_blocks": post_rows(args.api, "budget_rate_blocks", rate_blocks, args.dry_run),
        "budget_rate_materials": post_rows(args.api, "budget_rate_materials", rate_materials, args.dry_run),
    }
    print(json.dumps({
        "ok": True,
        "dryRun": args.dry_run,
        "counts": counts,
        "warnings": sorted(set(warnings))[:50],
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
