import json
import subprocess
import tempfile
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
MASTER_DIR = ROOT / "Master Data"
OUTPUT = ROOT / "webapp" / "data" / "master_data.json"


def scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, str):
        text = value.strip()
        return text if text else None
    return value


def clean_header(value: Any, fallback: str) -> str:
    text = str(scalar(value) or "").replace("\n", " ").strip()
    return text or fallback


def unique_headers(values: list[Any]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for idx, value in enumerate(values, start=1):
        name = clean_header(value, f"Column {idx}")
        seen[name] = seen.get(name, 0) + 1
        headers.append(name if seen[name] == 1 else f"{name} {seen[name]}")
    return headers


def is_empty(values: list[Any]) -> bool:
    return not any(scalar(v) not in (None, "") for v in values)


def detect_table(ws) -> tuple[int, list[str]]:
    sample_rows = []
    for row_no, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True), start=1):
        values = list(row)
        nonempty = sum(1 for v in values if scalar(v) not in (None, ""))
        sample_rows.append((row_no, values, nonempty))
        first = str(scalar(values[0]) or "").lower() if values else ""
        if first in {"column names", "colum name", "column name"}:
            return row_no, unique_headers(values)

    best = max(sample_rows, key=lambda x: (x[2], -x[0]), default=(1, [], 0))
    header_row = best[0]
    headers = unique_headers(best[1])
    while headers and headers[-1].startswith("Column "):
        headers.pop()
    return header_row, headers


def classify(file_name: str, sheet_name: str) -> str:
    text = f"{file_name} {sheet_name}".lower()
    if "partner" in text:
        return "คู่ค้า / Partner"
    if "fer" in text or "fert" in text or "ปุ๋ย" in text:
        return "ปุ๋ย / Fertilizer"
    if "ค่าแรง" in text or "payroll" in text or "wage" in text:
        return "ค่าแรง / Payroll"
    if "plant" in text or "terrain" in text or "block" in text or "summary_palm" in text:
        return "พื้นที่สวน / Plantation"
    if "cultivate" in text or "activity" in text or "material" in text or "crop" in text or "unit" in text:
        return "Cultivate Master"
    return "อื่น ๆ"


def convert_xls_to_xlsx(path: Path) -> Path | None:
    temp_dir = Path(tempfile.mkdtemp(prefix="master-data-xls-"))
    out = temp_dir / f"{path.stem}.xlsx"
    ps = f"""
$ErrorActionPreference='Stop'
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
$wb = $excel.Workbooks.Open('{str(path).replace("'", "''")}', 0, $true)
$wb.SaveAs('{str(out).replace("'", "''")}', 51)
$wb.Close($false)
$excel.Quit()
[System.Runtime.InteropServices.Marshal]::ReleaseComObject($wb) | Out-Null
[System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
"""
    try:
        subprocess.run(["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps], check=True, capture_output=True, text=True, timeout=120)
        return out if out.exists() else None
    except Exception:
        return None


def read_workbook(path: Path) -> dict[str, Any]:
    read_path = path
    warning = None
    if path.suffix.lower() == ".xls":
        converted = convert_xls_to_xlsx(path)
        if converted:
            read_path = converted
        else:
            warning = "ไม่สามารถแปลง .xls เป็น .xlsx ได้"

    info: dict[str, Any] = {
        "file": path.name,
        "path": str(path),
        "extension": path.suffix.lower(),
        "warning": warning,
        "sheets": [],
    }
    if warning:
        return info

    wb = openpyxl.load_workbook(read_path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        header_row, headers = detect_table(ws)
        if not headers:
            continue
        rows = []
        for offset, raw_values in enumerate(ws.iter_rows(min_row=header_row + 1, max_col=len(headers), values_only=True), start=header_row + 1):
            values = [scalar(value) for value in raw_values]
            if is_empty(values):
                continue
            row = {header: values[idx] if idx < len(values) else None for idx, header in enumerate(headers)}
            row["_id"] = f"{path.stem}::{ws.title}::{offset}"
            row["_rowNumber"] = offset
            rows.append(row)
        info["sheets"].append(
            {
                "name": ws.title,
                "group": classify(path.name, ws.title),
                "headerRow": header_row,
                "headers": headers,
                "rowCount": len(rows),
                "rows": rows,
            }
        )
    wb.close()
    return info


def main() -> None:
    files = []
    datasets = []
    for path in sorted(MASTER_DIR.rglob("*")):
        if not path.is_file() or path.name.startswith("~$") or path.suffix.lower() not in {".xlsx", ".xls"}:
            continue
        info = read_workbook(path)
        files.append({k: v for k, v in info.items() if k != "sheets"})
        for sheet in info.get("sheets", []):
            dataset_id = f"{path.stem}__{sheet['name']}".replace(" ", "_").replace("/", "_")
            datasets.append(
                {
                    "id": dataset_id,
                    "file": info["file"],
                    "sheet": sheet["name"],
                    "group": sheet["group"],
                    "headerRow": sheet["headerRow"],
                    "headers": sheet["headers"],
                    "rowCount": sheet["rowCount"],
                    "rows": sheet["rows"],
                }
            )

    groups = []
    for group in sorted({d["group"] for d in datasets}):
        group_sets = [d for d in datasets if d["group"] == group]
        groups.append(
            {
                "name": group,
                "datasetCount": len(group_sets),
                "rowCount": sum(d["rowCount"] for d in group_sets),
            }
        )

    payload = {
        "ok": True,
        "source": {
            "folder": str(MASTER_DIR),
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "fileCount": len(files),
            "datasetCount": len(datasets),
            "rowCount": sum(d["rowCount"] for d in datasets),
        },
        "groups": groups,
        "files": files,
        "datasets": datasets,
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"ok": True, "output": str(OUTPUT), "datasets": len(datasets), "rows": payload["source"]["rowCount"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
