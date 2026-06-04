import hashlib
import json
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
MASTER_DIR = ROOT / "Master Data"
OUT = ROOT / "webapp" / "data" / "master_data_full.json"
MAX_SCAN_ROWS = 3000
MAX_SCAN_COLS = 140


def clean(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value).strip()
    return value


def ascii_slug(text, fallback):
    raw = re.sub(r"[^a-zA-Z0-9]+", "_", text).strip("_").lower()
    return raw[:42] if raw else fallback


def unique_headers(values):
    headers = []
    seen = {}
    for index, value in enumerate(values, start=1):
        text = str(clean(value) or f"col_{index}")
        text = text[:80]
        count = seen.get(text, 0)
        seen[text] = count + 1
        headers.append(text if count == 0 else f"{text}_{count + 1}")
    return headers


def score_header(row):
    values = [clean(v) for v in row]
    non_blank = sum(1 for v in values if v != "")
    text_count = sum(1 for v in values if isinstance(v, str) and v.strip())
    code_words = sum(1 for v in values if isinstance(v, str) and re.search(r"code|รหัส|บล็อก|block|activity|partner|terrain|material", v, re.I))
    return non_blank + text_count * 2 + code_words * 3


def detect_header(rows):
    best_index = 0
    best_score = -1
    for i, row in enumerate(rows[:25]):
        score = score_header(row)
        if score > best_score:
            best_score = score
            best_index = i
    return best_index


def last_nonblank_col(rows):
    last = 0
    for row in rows:
        for idx, value in enumerate(row, start=1):
            if clean(value) != "":
                last = max(last, idx)
    return last


def infer_domain(file_name, sheet_name, columns):
    text = f"{file_name} {sheet_name} {' '.join(columns)}".lower()
    checks = [
        ("terrains", ["terrain", "บล็อก", "บล๊อก", "พื้นที่", "แปลง"]),
        ("activities", ["activity", "กิจกรรม", "workcode"]),
        ("partners", ["partner", "contractor", "employee", "first_name"]),
        ("materials", ["material", "fertilizer", "ปุ๋ย", "สารเคมี"]),
        ("equipment", ["equipment"]),
        ("warehouse", ["warehouse"]),
        ("budget", ["งบ", "ค่าใช้จ่าย", "ค่าแรง", "อัตราจ้าง", "ประมาณการ"]),
        ("crops", ["crop"]),
        ("companies", ["company"]),
        ("systems", ["system", "work system"]),
    ]
    for domain, needles in checks:
        if any(needle.lower() in text for needle in needles):
            return domain
    return "reference"


def infer_primary_key(columns, domain):
    priorities = [
        "block_code", "รหัสบล๊อกใหม่", "รหัสบล็อก", "รหัสบล๊อก", "Terrain given code", "terrain",
        "partner", "Material Given Code", "material_group", "Activity Code", "Activity given Code",
        "activity", "Company Code", "Crop given code", "WorkCode", "APcode", "code", "รหัส",
    ]
    lower_map = {str(col).lower(): col for col in columns}
    for name in priorities:
        for col in columns:
            if name.lower() == str(col).lower() or name.lower() in str(col).lower():
                return col
    if domain == "budget":
        for col in columns:
            if "บล็อก" in str(col) or "block" in str(col).lower():
                return col
    return columns[0] if columns else "id"


def infer_refs(columns):
    refs = []
    for col in columns:
        c = str(col).lower()
        if "block" in c or "บล็อก" in str(col) or "บล๊อก" in str(col) or "terrain" in c:
            refs.append({"field": col, "refDomain": "terrains", "refKey": "block_code/terrain"})
        elif "activity" in c or "กิจกรรม" in str(col) or "workcode" in c:
            refs.append({"field": col, "refDomain": "activities", "refKey": "activity_code"})
        elif "partner" in c or "contractor" in c or "employee" in c:
            refs.append({"field": col, "refDomain": "partners", "refKey": "partner"})
        elif "material" in c or "ปุ๋ย" in str(col) or "สาร" in str(col):
            refs.append({"field": col, "refDomain": "materials", "refKey": "material_code"})
        elif "company" in c:
            refs.append({"field": col, "refDomain": "companies", "refKey": "company_code"})
        elif "crop" in c:
            refs.append({"field": col, "refDomain": "crops", "refKey": "crop_code"})
    seen = set()
    out = []
    for ref in refs:
        key = (ref["field"], ref["refDomain"])
        if key not in seen:
            out.append(ref)
            seen.add(key)
    return out


CANONICAL_TABLES = [
    {
        "id": "master_terrains",
        "title": "ข้อมูลพื้นที่",
        "domain": "terrains",
        "primaryKey": "terrain_key",
        "primaryLabel": "terrain_key",
        "columns": [
            {"key": "terrain_key", "label": "คีย์พื้นที่", "type": "text"},
            {"key": "estate", "label": "Estate", "type": "text"},
            {"key": "area_group", "label": "แปลงใหญ่/พื้นที่", "type": "text"},
            {"key": "block_code", "label": "บล็อก", "type": "text"},
            {"key": "superior_key", "label": "พื้นที่แม่", "type": "text"},
            {"key": "rai", "label": "จำนวนไร่", "type": "number"},
            {"key": "tree_count", "label": "จำนวนต้น", "type": "number"},
            {"key": "planted_year", "label": "ปีปลูก", "type": "text"},
            {"key": "rspo", "label": "RSPO", "type": "text"},
            {"key": "status", "label": "สถานะ", "type": "text"},
        ],
        "references": [{"field": "superior_key", "refDomain": "terrains", "refKey": "terrain_key"}],
    },
    {
        "id": "master_activity_groups",
        "title": "กลุ่มกิจกรรม",
        "domain": "activities",
        "primaryKey": "activity_group_key",
        "primaryLabel": "activity_group_key",
        "columns": [
            {"key": "activity_group_key", "label": "คีย์กลุ่มกิจกรรม", "type": "text"},
            {"key": "activity_group_name", "label": "ชื่อกลุ่มกิจกรรม", "type": "text"},
            {"key": "status", "label": "สถานะ", "type": "text"},
        ],
        "references": [],
    },
    {
        "id": "master_activities",
        "title": "กิจกรรม",
        "domain": "activities",
        "primaryKey": "activity_key",
        "primaryLabel": "activity_key",
        "columns": [
            {"key": "activity_key", "label": "คีย์กิจกรรม", "type": "text"},
            {"key": "activity_name", "label": "ชื่อกิจกรรม", "type": "text"},
            {"key": "activity_group_key", "label": "คีย์กลุ่มกิจกรรม", "type": "text"},
            {"key": "unit", "label": "หน่วย", "type": "text"},
            {"key": "default_rate", "label": "อัตราตั้งต้น", "type": "number"},
            {"key": "status", "label": "สถานะ", "type": "text"},
        ],
        "references": [{"field": "activity_group_key", "refDomain": "activities", "refKey": "activity_group_key"}],
    },
    {
        "id": "master_partners",
        "title": "พนักงาน/ผู้รับเหมา/คู่ค้า",
        "domain": "partners",
        "primaryKey": "partner_key",
        "primaryLabel": "partner_key",
        "columns": [
            {"key": "partner_key", "label": "คีย์คู่ค้า", "type": "text"},
            {"key": "partner_name", "label": "ชื่อ", "type": "text"},
            {"key": "partner_type", "label": "ประเภท", "type": "text"},
            {"key": "company_key", "label": "คีย์บริษัท", "type": "text"},
            {"key": "estate", "label": "Estate", "type": "text"},
            {"key": "team", "label": "ทีม/กลุ่ม", "type": "text"},
            {"key": "wage_rate", "label": "ค่าแรง", "type": "number"},
            {"key": "status", "label": "สถานะ", "type": "text"},
        ],
        "references": [{"field": "company_key", "refDomain": "companies", "refKey": "company_key"}],
    },
    {
        "id": "master_materials",
        "title": "วัสดุ/ปุ๋ย/สารเคมี",
        "domain": "materials",
        "primaryKey": "material_key",
        "primaryLabel": "material_key",
        "columns": [
            {"key": "material_key", "label": "คีย์วัสดุ", "type": "text"},
            {"key": "material_name", "label": "ชื่อวัสดุ", "type": "text"},
            {"key": "material_group_key", "label": "คีย์กลุ่มวัสดุ", "type": "text"},
            {"key": "unit", "label": "หน่วย", "type": "text"},
            {"key": "standard_rate", "label": "อัตรามาตรฐาน", "type": "number"},
            {"key": "status", "label": "สถานะ", "type": "text"},
        ],
        "references": [{"field": "material_group_key", "refDomain": "materials", "refKey": "material_group_key"}],
    },
    {
        "id": "master_budget_lines",
        "title": "งบประมาณกิจกรรมรายบล็อก",
        "domain": "budget",
        "primaryKey": "budget_key",
        "primaryLabel": "budget_key",
        "columns": [
            {"key": "budget_key", "label": "คีย์งบประมาณ", "type": "text"},
            {"key": "fiscal_year", "label": "ปีงบประมาณ", "type": "text"},
            {"key": "terrain_key", "label": "คีย์พื้นที่", "type": "text"},
            {"key": "activity_key", "label": "คีย์กิจกรรม", "type": "text"},
            {"key": "rai", "label": "จำนวนไร่", "type": "number"},
            {"key": "tree_count", "label": "จำนวนต้น", "type": "number"},
            {"key": "rate", "label": "อัตรา", "type": "number"},
            {"key": "budget_amount", "label": "งบประมาณ", "type": "number"},
        ],
        "references": [
            {"field": "terrain_key", "refDomain": "terrains", "refKey": "terrain_key"},
            {"field": "activity_key", "refDomain": "activities", "refKey": "activity_key"},
        ],
    },
    {
        "id": "master_equipment",
        "title": "เครื่องจักร/อุปกรณ์",
        "domain": "equipment",
        "primaryKey": "equipment_key",
        "primaryLabel": "equipment_key",
        "columns": [
            {"key": "equipment_key", "label": "คีย์เครื่องจักร", "type": "text"},
            {"key": "equipment_name", "label": "ชื่อเครื่องจักร", "type": "text"},
            {"key": "equipment_type", "label": "ประเภท", "type": "text"},
            {"key": "terrain_key", "label": "คีย์พื้นที่", "type": "text"},
            {"key": "status", "label": "สถานะ", "type": "text"},
        ],
        "references": [{"field": "terrain_key", "refDomain": "terrains", "refKey": "terrain_key"}],
    },
    {
        "id": "master_warehouses",
        "title": "คลัง/สถานที่จัดเก็บ",
        "domain": "warehouse",
        "primaryKey": "warehouse_key",
        "primaryLabel": "warehouse_key",
        "columns": [
            {"key": "warehouse_key", "label": "คีย์คลัง", "type": "text"},
            {"key": "warehouse_name", "label": "ชื่อคลัง", "type": "text"},
            {"key": "company_key", "label": "คีย์บริษัท", "type": "text"},
            {"key": "status", "label": "สถานะ", "type": "text"},
        ],
        "references": [{"field": "company_key", "refDomain": "companies", "refKey": "company_key"}],
    },
]


def first_value(row, names):
    for name in names:
        for key, value in row.items():
            if key.startswith("_"):
                continue
            if name.lower() in str(key).lower() and clean(value) != "":
                return clean(value)
    return ""


def add_unique(target, seen, item, key_name):
    value = str(item.get(key_name, "")).strip()
    if not value or value in seen:
        return
    seen.add(value)
    target.append(item)


def build_canonical_tables(raw_tables):
    configs = {table["id"]: {**table, "rows": [], "rowCount": 0, "columnCount": len(table["columns"])} for table in CANONICAL_TABLES}
    seen = {key: set() for key in configs}
    activity_groups = {}
    for table in raw_tables:
        for row in table["rows"]:
            if table["domain"] == "terrains":
                terrain = first_value(row, ["รหัสบล๊อกใหม่", "รหัสบล็อก", "รหัสบล๊อก", "terrain given code", "terrain", "บล็อก", "บล๊อก"])
                block = first_value(row, ["บล็อก", "บล๊อก", "terrain"])
                add_unique(configs["master_terrains"]["rows"], seen["master_terrains"], {
                    "terrain_key": terrain or block,
                    "estate": first_value(row, ["estate", "company"]),
                    "area_group": first_value(row, ["แปลง", "พื้นที่", "wparea", "area"]),
                    "block_code": block or terrain,
                    "superior_key": first_value(row, ["superior", "พื้นที่แม่"]),
                    "rai": first_value(row, ["จน.ไร่", "จำนวนไร่", "terrain area", "area planted", "area"]),
                    "tree_count": first_value(row, ["จน.ต้น", "# of trees", "tree count", "จำนวนต้น"]),
                    "planted_year": first_value(row, ["ปีปลูก", "date planted", "อายุ", "รุ่น"]),
                    "rspo": first_value(row, ["rspo"]),
                    "status": first_value(row, ["status"]) or "active",
                }, "terrain_key")
            elif table["domain"] == "activities":
                group_key = first_value(row, ["activity group given code", "activity group code", "activity_group"])
                group_name = first_value(row, ["activity group name", "activity group"])
                if group_key or group_name:
                    activity_groups[group_key or group_name] = group_name or group_key
                    add_unique(configs["master_activity_groups"]["rows"], seen["master_activity_groups"], {
                        "activity_group_key": group_key or group_name,
                        "activity_group_name": group_name or group_key,
                        "status": "active",
                    }, "activity_group_key")
                activity_key = first_value(row, ["activity given code", "activity code", "activity", "workcode"])
                activity_name = first_value(row, ["activity name", "activity", "workname", "description", "กิจกรรม"])
                add_unique(configs["master_activities"]["rows"], seen["master_activities"], {
                    "activity_key": activity_key,
                    "activity_name": activity_name or activity_key,
                    "activity_group_key": group_key,
                    "unit": first_value(row, ["uom", "unit", "หน่วย"]),
                    "default_rate": first_value(row, ["rate", "อัตรา"]),
                    "status": first_value(row, ["status"]) or "active",
                }, "activity_key")
            elif table["domain"] == "partners":
                partner = first_value(row, ["partner"])
                add_unique(configs["master_partners"]["rows"], seen["master_partners"], {
                    "partner_key": partner,
                    "partner_name": first_value(row, ["description", "first_name", "name"]),
                    "partner_type": first_value(row, ["partner_type", "designation", "type"]),
                    "company_key": first_value(row, ["company"]),
                    "estate": first_value(row, ["estate"]),
                    "team": first_value(row, ["gang", "team"]),
                    "wage_rate": first_value(row, ["rate", "ค่าแรง"]),
                    "status": first_value(row, ["status"]) or "active",
                }, "partner_key")
            elif table["domain"] == "materials":
                material = first_value(row, ["material given code", "material", "ปุ๋ย"])
                group = first_value(row, ["material group", "material_group"])
                add_unique(configs["master_materials"]["rows"], seen["master_materials"], {
                    "material_key": material or group,
                    "material_name": first_value(row, ["material name", "name", "description"]) or material,
                    "material_group_key": group,
                    "unit": first_value(row, ["uom", "unit", "measurement"]),
                    "standard_rate": first_value(row, ["rate", "dosage", "อัตรา"]),
                    "status": first_value(row, ["status"]) or "active",
                }, "material_key")
            elif table["domain"] == "budget":
                terrain = first_value(row, ["บล็อก", "block", "terrain"])
                activity = table["title"]
                budget_key = f"2569-{terrain}-{activity}-{row.get('_sourceRow', '')}"
                add_unique(configs["master_budget_lines"]["rows"], seen["master_budget_lines"], {
                    "budget_key": budget_key,
                    "fiscal_year": "2569",
                    "terrain_key": terrain,
                    "activity_key": activity,
                    "rai": first_value(row, ["ไร่", "จำนวนไร่"]),
                    "tree_count": first_value(row, ["ต้น", "จำนวนต้น"]),
                    "rate": first_value(row, ["ค่าแรง", "อัตรา", "rate"]),
                    "budget_amount": first_value(row, ["รวม", "งบ", "บาท"]),
                }, "budget_key")
            elif table["domain"] == "equipment":
                equipment = first_value(row, ["equipment given code", "equipment"])
                add_unique(configs["master_equipment"]["rows"], seen["master_equipment"], {
                    "equipment_key": equipment,
                    "equipment_name": first_value(row, ["equipment name", "name"]),
                    "equipment_type": first_value(row, ["equipment type"]),
                    "terrain_key": first_value(row, ["estate code", "terrain"]),
                    "status": first_value(row, ["status"]) or "active",
                }, "equipment_key")
            elif table["domain"] == "warehouse":
                warehouse = first_value(row, ["warehouseactivity given code", "warehouse"])
                add_unique(configs["master_warehouses"]["rows"], seen["master_warehouses"], {
                    "warehouse_key": warehouse,
                    "warehouse_name": first_value(row, ["warehouse name", "name"]),
                    "company_key": first_value(row, ["company code", "company"]),
                    "status": "active",
                }, "warehouse_key")
    for table in configs.values():
        for index, row in enumerate(table["rows"], start=1):
            row["_sourceRow"] = index
        table["rowCount"] = len(table["rows"])
    return list(configs.values())


def sheet_table(path, ws):
    raw_rows = []
    max_col = min(ws.max_column or 0, MAX_SCAN_COLS)
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, MAX_SCAN_ROWS), max_col=max_col, values_only=True):
        values = [clean(v) for v in row]
        if any(v != "" for v in values):
            raw_rows.append(values)
    if not raw_rows:
        return None
    header_index = detect_header(raw_rows)
    width = max(last_nonblank_col(raw_rows[: min(len(raw_rows), header_index + 30)]), 1)
    headers = unique_headers(raw_rows[header_index][:width])
    rows = []
    for idx, raw in enumerate(raw_rows[header_index + 1 :], start=header_index + 2):
        item = {}
        for i, header in enumerate(headers):
            item[header] = clean(raw[i]) if i < len(raw) else ""
        if any(v != "" for v in item.values()):
            item["_sourceRow"] = idx
            rows.append(item)
    seed = f"{path.name}|{ws.title}".encode("utf-8")
    digest = hashlib.sha1(seed).hexdigest()[:8]
    domain = infer_domain(path.name, ws.title, headers)
    table_id = f"md_{ascii_slug(path.stem, 'file')}_{ascii_slug(ws.title, 'sheet')}_{digest}"
    primary_key = infer_primary_key(headers, domain)
    return {
        "id": table_id,
        "title": ws.title,
        "domain": domain,
        "sourceFile": path.name,
        "sourcePath": str(path),
        "headerRow": header_index + 1,
        "rowCount": len(rows),
        "columnCount": len(headers),
        "primaryKey": primary_key,
        "primaryLabel": str(primary_key),
        "references": infer_refs(headers),
        "columns": [{"key": h, "label": h, "type": "text"} for h in headers],
        "rows": rows,
    }


def main():
    tables = []
    skipped = []
    for path in sorted(MASTER_DIR.rglob("*")):
        if not path.is_file() or path.name.startswith("~$"):
            continue
        if path.suffix.lower() == ".xlsx":
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            for sheet in wb.sheetnames:
                table = sheet_table(path, wb[sheet])
                if table:
                    tables.append(table)
        elif path.suffix.lower() == ".xls":
            skipped.append({"file": path.name, "reason": "legacy .xls; convert to .xlsx to import rows"})
    domains = {}
    for table in tables:
        domains.setdefault(table["domain"], {"id": table["domain"], "tableCount": 0, "rowCount": 0})
        domains[table["domain"]]["tableCount"] += 1
        domains[table["domain"]]["rowCount"] += table["rowCount"]
    canonical_tables = build_canonical_tables(tables)
    canonical_domains = {}
    for table in canonical_tables:
        canonical_domains.setdefault(table["domain"], {"id": table["domain"], "tableCount": 0, "rowCount": 0})
        canonical_domains[table["domain"]]["tableCount"] += 1
        canonical_domains[table["domain"]]["rowCount"] += table["rowCount"]
    OUT.write_text(json.dumps({
        "ok": True,
        "sourceFolder": str(MASTER_DIR),
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "tableCount": len(canonical_tables),
        "rowCount": sum(t["rowCount"] for t in canonical_tables),
        "domains": list(canonical_domains.values()),
        "tables": canonical_tables,
        "rawTableCount": len(tables),
        "rawRowCount": sum(t["rowCount"] for t in tables),
        "rawDomains": list(domains.values()),
        "skipped": skipped,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {OUT} tables={len(canonical_tables)} rows={sum(t['rowCount'] for t in canonical_tables)} raw_tables={len(tables)} skipped={len(skipped)}")


if __name__ == "__main__":
    main()
