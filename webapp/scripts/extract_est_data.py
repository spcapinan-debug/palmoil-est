import json
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import openpyxl
from docx import Document


ROOT = Path(__file__).resolve().parents[2]
MASTER_DIR = ROOT / "Master Data"
BUDGET_FILE = MASTER_DIR / "ประมาณการค่าใช้จ่าย 2569.xlsx"
EST_FILE = MASTER_DIR / "est.docx"
MASTER_DATA_FILE = ROOT / "webapp" / "data" / "master_data.json"
OUTPUT = ROOT / "webapp" / "data" / "est_data.json"


def value_of(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, str):
        text = value.strip()
        return text if text else None
    return value


def n(value: Any) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(str(value).replace(",", ""))
    except Exception:
        return 0.0


def clean_header(value: Any, fallback: str) -> str:
    text = str(value_of(value) or "").replace("\n", " ").strip()
    return text or fallback


def unique_headers(values: list[Any]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for idx, raw in enumerate(values, start=1):
        name = clean_header(raw, f"Column {idx}")
        seen[name] = seen.get(name, 0) + 1
        headers.append(name if seen[name] == 1 else f"{name} {seen[name]}")
    while headers and headers[-1].startswith("Column "):
        headers.pop()
    return headers


def detect_header(ws) -> tuple[int, list[str]]:
    best = (1, [], 0)
    max_scan = min(ws.max_row, 12)
    max_col = min(ws.max_column, 60)
    for row_no, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, max_col=max_col, values_only=True), start=1):
        values = list(row)
        nonempty = sum(1 for v in values if value_of(v) not in (None, ""))
        text = " ".join(str(value_of(v) or "") for v in values[:12])
        score = nonempty
        if any(key in text for key in ["บล็อก", "แปลง", "ไร่", "ต้น", "ค่าแรง", "ผลผลิต", "รหัสบล็อก"]):
            score += 6
        if score > best[2]:
            best = (row_no, values, score)
    return best[0], unique_headers(best[1])


def classify_activity(sheet_name: str) -> str:
    text = sheet_name.lower()
    if "ผลผลิต" in text:
        return "ตัดปาล์ม / ผลผลิต"
    if "ปุ๋ย" in text or "fer" in text:
        return "ใส่ปุ๋ย"
    if "สารเคมี" in text or "ถาง" in text or "ถาก" in text or "ตัดหญ้า" in text or "แหวก" in text:
        return "กำจัดวัชพืช"
    if "ฟีโรโมน" in text or "พิโพรนิล" in text or "เจาะ" in text:
        return "กำจัดศัตรูพืช"
    if "ไถ" in text or "พรวน" in text:
        return "เตรียมพื้นที่"
    if "ปาล์ม" in text or "ต้น" in text:
        return "ข้อมูลพื้นที่"
    return "งานอื่น ๆ"


def is_total_row(row: dict[str, Any]) -> bool:
    text = " ".join(str(v or "") for k, v in row.items() if not k.startswith("_"))
    return "รวม" in text or "total" in text.lower()


def extract_budget() -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(BUDGET_FILE, read_only=True, data_only=True)
    datasets = []
    for ws in wb.worksheets:
        header_row, headers = detect_header(ws)
        rows = []
        totals = {
            "areaRai": 0.0,
            "trees": 0.0,
            "productionTon": 0.0,
            "laborCost": 0.0,
            "materialQty": 0.0,
        }
        for row_no, raw in enumerate(ws.iter_rows(min_row=header_row + 1, max_col=len(headers), values_only=True), start=header_row + 1):
            values = [value_of(v) for v in raw]
            if not any(v not in (None, "") for v in values):
                continue
            row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
            row["_id"] = f"{ws.title}::{row_no}"
            row["_rowNumber"] = row_no
            row["_activity"] = classify_activity(ws.title)
            row["_isTotal"] = is_total_row(row)
            rows.append(row)
            if row["_isTotal"]:
                continue
            for key, value in row.items():
                label = str(key)
                if "ไร่" in label:
                    totals["areaRai"] += n(value)
                if "ต้น" in label:
                    totals["trees"] += n(value)
                if "ผลผลิต" in label or "ตัน" in label:
                    totals["productionTon"] += n(value)
                if "ค่า" in label or "บาท" in label or "รวม" in label:
                    totals["laborCost"] += n(value)
                if "ปริมาณ" in label or "ลิตร" in label or "กิโล" in label or "kg" in label.lower():
                    totals["materialQty"] += n(value)
        datasets.append(
            {
                "id": ws.title.replace(" ", "_"),
                "sheet": ws.title,
                "activity": classify_activity(ws.title),
                "headerRow": header_row,
                "headers": headers,
                "rowCount": len(rows),
                "totals": totals,
                "rows": rows,
            }
        )
    wb.close()
    return datasets


def extract_est_doc() -> dict[str, Any]:
    doc = Document(EST_FILE)
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    menus = []
    current = None
    for text in paragraphs:
        if text.startswith("เมนู") or text.startswith("ชื่อระบบ"):
            current = {"title": text, "items": []}
            menus.append(current)
        elif current:
            current["items"].append(text)
    return {"paragraphs": paragraphs, "menus": menus}


def load_master_summary() -> dict[str, Any]:
    if not MASTER_DATA_FILE.exists():
        return {"groups": [], "datasets": [], "source": {}}
    payload = json.loads(MASTER_DATA_FILE.read_text(encoding="utf-8"))
    return {
        "source": payload.get("source", {}),
        "groups": payload.get("groups", []),
        "datasets": [
            {
                "id": item.get("id"),
                "file": item.get("file"),
                "sheet": item.get("sheet"),
                "group": item.get("group"),
                "rowCount": item.get("rowCount", 0),
                "headers": item.get("headers", [])[:12],
            }
            for item in payload.get("datasets", [])
        ],
    }


def main() -> None:
    budget = extract_budget()
    est_doc = extract_est_doc()
    master_summary = load_master_summary()
    activity_totals = {}
    for dataset in budget:
        bucket = activity_totals.setdefault(dataset["activity"], {"datasets": 0, "rows": 0, "laborCost": 0.0, "areaRai": 0.0, "trees": 0.0})
        bucket["datasets"] += 1
        bucket["rows"] += dataset["rowCount"]
        bucket["laborCost"] += dataset["totals"]["laborCost"]
        bucket["areaRai"] += dataset["totals"]["areaRai"]
        bucket["trees"] += dataset["totals"]["trees"]
    payload = {
        "ok": True,
        "source": {
            "budgetFile": BUDGET_FILE.name,
            "estFile": EST_FILE.name,
            "masterFolder": str(MASTER_DIR),
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "datasetCount": len(budget),
            "rowCount": sum(item["rowCount"] for item in budget),
        },
        "menu": [
            {"id": "est-dashboard", "title": "Dashboard", "group": "ระบบการจัดการสวนปาล์ม", "source": "est.docx"},
            {"id": "est-master", "title": "ข้อมูลหลัก", "group": "ข้อมูลหลัก", "source": "Master Data"},
            {"id": "est-budget", "title": "งบประมาณ 2569", "group": "งบประมาณ", "source": BUDGET_FILE.name},
            {"id": "est-plan", "title": "วางแผนงาน", "group": "Plan", "source": "est.docx"},
            {"id": "est-workorder", "title": "สั่งงาน", "group": "Work Order", "source": "est.docx"},
            {"id": "est-daily", "title": "บันทึกทำงาน", "group": "Daily Work", "source": "est.docx"},
            {"id": "est-payroll", "title": "อัตราค่าแรง", "group": "Payroll", "source": "est.docx"},
            {"id": "est-report", "title": "รายงาน", "group": "Reports", "source": "est.docx"},
            {"id": "est-stack", "title": "Vercel + Supabase + GitHub", "group": "System", "source": "Stack target"},
        ],
        "estDoc": est_doc,
        "activityTotals": activity_totals,
        "budgetDatasets": budget,
        "masterSummary": master_summary,
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"ok": True, "output": str(OUTPUT), "datasets": len(budget), "rows": payload["source"]["rowCount"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
