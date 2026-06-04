from __future__ import annotations

import json
import argparse
import subprocess
import tempfile
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = ROOT / "Summary_Palm_RSPO-Ramp.xlsx"
OUTPUT = Path(__file__).resolve().parents[1] / "data" / "data.json"
QUERY_SCRIPT = Path(__file__).resolve().parent / "query_weight_data.ps1"

REPORT_SHEETS = {
    "Movement": {"header": 7, "start": 8, "cols": 16, "date_col": 1},
    "Daily Pivot": {"header": 4, "start": 5, "cols": 18, "date_col": 1},
    "Stock Report": {"header": 6, "start": 7, "cols": 31, "date_col": 1},
    "Summary": {"header": 4, "start": 5, "cols": 18, "date_col": 1},
    "RSPO Report": {"header": 5, "start": 6, "cols": 18, "date_col": 5},
    "Pivot Movement": {"header": 4, "start": 5, "cols": 18, "date_col": None},
    "Clear_Ramp_Log": {"header": 1, "start": 2, "cols": 21, "date_col": 1},
}

THAI_MONTHS = {
    "มกราคม": 1,
    "กุมภาพันธ์": 2,
    "มีนาคม": 3,
    "เมษายน": 4,
    "พฤษภาคม": 5,
    "มิถุนายน": 6,
    "กรกฎาคม": 7,
    "สิงหาคม": 8,
    "กันยายน": 9,
    "ตุลาคม": 10,
    "พฤศจิกายน": 11,
    "ธันวาคม": 12,
}

DATA_FIELDS = [
    "wpDocNo",
    "wpInOutType",
    "wpDocDate",
    "wptypecar",
    "wpctCode",
    "wpCarLicense",
    "wpCarProvince",
    "wpDriver",
    "wpGrade",
    "wpHeadCode",
    "wpProduct",
    "wpBunchPrice",
    "wpCarWeightDate",
    "wpCarWeight",
    "wpTotalWeightDate",
    "wpTotalWeight",
    "wpNetWeight",
    "wpRipePercent",
    "wpRawPercent",
    "wpFallPercent",
    "wpMixRawPercent",
    "wpBadPercent",
    "wpDuraPercent",
    "wpOtherPercent",
    "wpLongStemPercent",
    "wpRampAmount",
    "wpReturnFFB",
    "wpFacDocNo",
    "wpFacDocDate",
    "wpFacGrade",
    "wpFacCarWeight",
    "wpFacTotalWeight",
    "wpFacNetWeight",
    "wpReturnPalm",
    "wpftcode",
    "wpfacprice",
    "wpintime",
    "wpouttime",
    "location",
]


def excel_serial_to_iso(value: int | float) -> str:
    base = datetime(1899, 12, 30)
    return (base + timedelta(days=float(value))).isoformat(timespec="seconds")


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, str):
        text = value.strip()
        return text if text else None
    return value


def number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def day(value: Any) -> str | None:
    value = clean_value(value)
    if not value:
        return None
    return str(value)[:10]


def read_lookup_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict[str, Any]]:
    headers = [clean_value(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        if not any(v not in (None, "") for v in values):
            continue
        rows.append({str(headers[i]).strip(): clean_value(values[i]) for i in range(min(len(headers), len(values))) if headers[i]})
    return rows


def build_lookups(wb: openpyxl.Workbook) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    terrain_rows = read_lookup_sheet(wb["Terrain"])
    wpct_rows = read_lookup_sheet(wb["Wpct"])

    terrain_by_code = {}
    for row in terrain_rows:
        code = clean_value(row.get("WpctCode") or row.get("LookupKey_fast"))
        if code is not None:
            terrain_by_code[str(code)] = row

    wpct_by_code = {}
    for row in wpct_rows:
        code = clean_value(row.get("WpctCode") or row.get("WpctCode ") or row.get("LookupKey_fast"))
        if code is not None:
            wpct_by_code[str(code)] = row

    return terrain_by_code, wpct_by_code, terrain_rows, wpct_rows


def classify(row: dict[str, Any], terrain: dict[str, Any] | None) -> str:
    if row.get("wpInOutType") == "O":
        return "Outbound Logistics"
    if terrain and str(terrain.get("RSPO Certified", "")).upper() == "YES":
        return "RSPO"
    code = str(row.get("wpctCode") or "")
    if code.startswith("22"):
        return "Contract Farmer"
    return "NON-RSPO"


def read_records(wb: openpyxl.Workbook, terrain_by_code: dict[str, dict[str, Any]], wpct_by_code: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    ws = wb["Data"]
    headers = [clean_value(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {header: pos for pos, header in enumerate(headers)}
    rows = []

    for source_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v not in (None, "") for v in values):
            continue

        item = {"_srcRow": source_row}
        for field in DATA_FIELDS:
            raw = values[index[field]] if field in index and index[field] < len(values) else None
            value = clean_value(raw)
            if field in {"wpintime", "wpouttime"} and isinstance(raw, (int, float)) and raw > 1:
                value = excel_serial_to_iso(raw)
            item[field] = value

        code = str(item.get("wpctCode") or "")
        terrain = terrain_by_code.get(code)
        wpct = wpct_by_code.get(code)
        item["date"] = day(item.get("wpDocDate"))
        item["weightDate"] = day(item.get("wpCarWeightDate")) or item["date"]
        item["name"] = clean_value((wpct or {}).get("WpctName")) or clean_value((terrain or {}).get("Payroll Code Description")) or code
        item["standard"] = classify(item, terrain)
        item["estate"] = clean_value((terrain or {}).get("estate"))
        item["superior"] = clean_value((terrain or {}).get("superior"))
        item["areaGroup"] = clean_value((terrain or {}).get("wparea"))
        item["terrain"] = clean_value((terrain or {}).get("terrain"))
        item["palmSize"] = clean_value((terrain or {}).get("ขนาดปาล์ม"))
        item["rspoCertified"] = clean_value((terrain or {}).get("RSPO Certified"))
        rows.append(item)

    return rows


def build_record_from_source(source_row: int, data: dict[str, Any], terrain_by_code: dict[str, dict[str, Any]], wpct_by_code: dict[str, dict[str, Any]]) -> dict[str, Any]:
    item = {"_srcRow": source_row}
    for field in DATA_FIELDS:
        item[field] = clean_value(data.get(field))
    code = str(item.get("wpctCode") or "")
    terrain = terrain_by_code.get(code)
    wpct = wpct_by_code.get(code)
    item["date"] = day(item.get("wpDocDate"))
    item["weightDate"] = day(item.get("wpCarWeightDate")) or item["date"]
    item["name"] = clean_value((wpct or {}).get("WpctName")) or clean_value((terrain or {}).get("Payroll Code Description")) or code
    item["standard"] = classify(item, terrain)
    item["estate"] = clean_value((terrain or {}).get("estate"))
    item["superior"] = clean_value((terrain or {}).get("superior"))
    item["areaGroup"] = clean_value((terrain or {}).get("wparea"))
    item["terrain"] = clean_value((terrain or {}).get("terrain"))
    item["palmSize"] = clean_value((terrain or {}).get("เธเธเธฒเธ”เธเธฒเธฅเนเธก"))
    item["rspoCertified"] = clean_value((terrain or {}).get("RSPO Certified"))
    return item


def read_records_from_query(terrain_by_code: dict[str, dict[str, Any]], wpct_by_code: dict[str, dict[str, Any]]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if not QUERY_SCRIPT.is_file():
        raise FileNotFoundError(f"Missing query script: {QUERY_SCRIPT}")
    with tempfile.TemporaryDirectory(prefix="rspo-query-") as tmp:
        raw_output = Path(tmp) / "weight_query_rows.json"
        cmd = [
            "powershell",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(QUERY_SCRIPT),
            "-Workbook",
            str(WORKBOOK),
            "-Output",
            str(raw_output),
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=240)
        if result.returncode != 0:
            message = (result.stderr or result.stdout or "ODBC query failed").strip()
            raise RuntimeError(message)
        payload = json.loads(raw_output.read_text(encoding="utf-8-sig"))
    raw_rows = payload.get("rows") or []
    records = [
        build_record_from_source(index, row, terrain_by_code, wpct_by_code)
        for index, row in enumerate(raw_rows, start=2)
    ]
    return records, payload.get("source") or {}


def make_unique_headers(headers: list[Any]) -> list[str]:
    seen: dict[str, int] = {}
    out = []
    for i, header in enumerate(headers, start=1):
        text = str(clean_value(header) or f"Column {i}")
        count = seen.get(text, 0) + 1
        seen[text] = count
        out.append(text if count == 1 else f"{text} ({count})")
    return out


def read_report_sheet(wb: openpyxl.Workbook, name: str, spec: dict[str, Any]) -> dict[str, Any]:
    ws = wb[name]
    cols = min(spec["cols"], ws.max_column)
    header_values = next(ws.iter_rows(min_row=spec["header"], max_row=spec["header"], min_col=1, max_col=cols, values_only=True))
    headers = make_unique_headers(list(header_values))
    rows = []

    for offset, values in enumerate(ws.iter_rows(min_row=spec["start"], max_row=ws.max_row, min_col=1, max_col=cols, values_only=True), start=spec["start"]):
        values = [clean_value(v) for v in values]
        if not any(v not in (None, "") for v in values):
            continue
        row = {headers[i]: values[i] for i in range(cols)}
        row["_row"] = offset
        date_col = spec.get("date_col")
        row["_date"] = day(values[date_col - 1]) if date_col else None
        rows.append(row)

    return {
        "title": clean_value(ws.cell(1, 1).value) or name,
        "subtitle": clean_value(ws.cell(2, 1).value),
        "headers": headers,
        "rows": rows,
    }


def add_to_bucket(bucket: dict[str, Any], row: dict[str, Any]) -> None:
    net = number(row.get("wpNetWeight"))
    fac = number(row.get("wpFacNetWeight"))
    ramp = number(row.get("wpRampAmount"))
    is_in = row.get("wpInOutType") == "I"
    is_out = row.get("wpInOutType") == "O"

    bucket["rows"] += 1
    bucket["net"] += net
    bucket["facNet"] += fac
    bucket["ramp"] += ramp
    if is_in:
        bucket["in"] += net
        bucket["inRows"] += 1
    if is_out:
        bucket["out"] += net
        bucket["outRows"] += 1
        bucket["diff"] += fac - net


def empty_bucket(key: str) -> dict[str, Any]:
    return {"key": key, "rows": 0, "inRows": 0, "outRows": 0, "net": 0, "in": 0, "out": 0, "facNet": 0, "ramp": 0, "diff": 0}


def build_analytics(records: list[dict[str, Any]]) -> dict[str, Any]:
    by_day = defaultdict(lambda: empty_bucket(""))
    by_standard = defaultdict(lambda: empty_bucket(""))
    by_area = defaultdict(lambda: empty_bucket(""))
    by_name = defaultdict(lambda: empty_bucket(""))

    for row in records:
        for group, key in [
            (by_day, row.get("date") or "(no date)"),
            (by_standard, row.get("standard") or "(no standard)"),
            (by_area, row.get("areaGroup") or "Unknown"),
            (by_name, row.get("name") or row.get("wpctCode") or "Unknown"),
        ]:
            group[key]["key"] = key
            add_to_bucket(group[key], row)

    def sorted_values(mapping: dict[str, dict[str, Any]], key: str = "net") -> list[dict[str, Any]]:
        return sorted(mapping.values(), key=lambda item: item[key], reverse=True)

    daily = sorted(by_day.values(), key=lambda item: item["key"])
    return {
        "daily": daily,
        "standard": sorted_values(by_standard),
        "area": sorted_values(by_area),
        "topNames": sorted_values(by_name)[:40],
    }


def month_from_filename(name: str) -> tuple[int, int] | None:
    for month_name, month in THAI_MONTHS.items():
        if month_name in name:
            return 2026, month
    return None


def read_monthly_stock_reports() -> dict[str, Any]:
    reports = {
        "garden": {"headers": [], "rows": []},
        "takuk": {"headers": [], "rows": []},
        "combined": {"headers": [], "rows": []},
    }
    sheet_keys = ["garden", "takuk", "combined"]

    for path in ROOT.glob("*.xlsx"):
        if not path.name.startswith("สรุปสต๊อคผลปาล์มสดประจำเดือน"):
            continue
        month_info = month_from_filename(path.name)
        if not month_info:
            continue
        year, month = month_info
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for key, ws in zip(sheet_keys, wb.worksheets[:3]):
            max_col = ws.max_column
            header_rows = list(ws.iter_rows(min_row=5, max_row=6, min_col=1, max_col=max_col, values_only=True))
            header1 = [clean_value(value) for value in header_rows[0]]
            header2 = [clean_value(value) for value in header_rows[1]]
            if not reports[key]["headers"] or month == 1:
                reports[key]["headers"] = [header1, header2]

            for values in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=1, max_col=max_col, values_only=True):
                day_value = values[0]
                if not isinstance(day_value, (int, float)) or not 1 <= int(day_value) <= 31:
                    continue
                try:
                    report_date = date(year, month, int(day_value)).isoformat()
                except ValueError:
                    continue
                cells = [clean_value(value) for value in values]
                reports[key]["rows"].append({
                    "date": report_date,
                    "monthFile": path.name,
                    "sheet": ws.title,
                    "cells": cells,
                })

    for report in reports.values():
        report["rows"].sort(key=lambda row: row["date"])
    return reports


def read_workbook_stock_report(wb: openpyxl.Workbook) -> dict[str, Any]:
    ws = wb["Stock Report"]
    max_col = 33
    header_rows = list(ws.iter_rows(min_row=5, max_row=6, min_col=1, max_col=max_col, values_only=True))
    header1 = [clean_value(value) for value in header_rows[0]]
    header2 = [clean_value(value) for value in header_rows[1]]
    rows = []
    for values in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=1, max_col=max_col, values_only=True):
        report_date = day(values[0])
        if not report_date:
            continue
        cells = [clean_value(value) for value in values]
        if not any(value not in (None, "") for value in cells[1:]):
            continue
        # Match the monthly "รวม" report shape: hide helper columns after note.
        rows.append({
            "date": report_date,
            "monthFile": WORKBOOK.name,
            "sheet": "Stock Report",
            "cells": cells,
        })
    return {
        "combined": {
            "headers": [header1, header2],
            "rows": rows,
        }
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", choices=["sheet", "query"], default="query")
    args = parser.parse_args()

    wb_values = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    terrain_by_code, wpct_by_code, terrain_rows, wpct_rows = build_lookups(wb_values)
    query_source = {}
    if args.source == "query":
        records, query_source = read_records_from_query(terrain_by_code, wpct_by_code)
    else:
        records = read_records(wb_values, terrain_by_code, wpct_by_code)

    date_values = [r["date"] for r in records if r.get("date")]
    sheet_tables = {name: read_report_sheet(wb_values, name, spec) for name, spec in REPORT_SHEETS.items()}

    payload = {
        "source": {
            "workbook": WORKBOOK.name,
            "recordSource": args.source,
            "query": query_source,
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "rowCount": len(records),
            "dateMin": min(date_values) if date_values else None,
            "dateMax": max(date_values) if date_values else None,
            "linkedFrom": ["Data", "Terrain", "Wpct"],
            "views": list(REPORT_SHEETS),
        },
        "records": records,
        "analytics": build_analytics(records),
        "sheets": sheet_tables,
        "lookups": {
            "terrain": terrain_rows,
            "wpct": wpct_rows,
        },
        "monthlyReports": read_monthly_stock_reports(),
        "workbookReports": read_workbook_stock_report(wb_values),
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"Wrote {len(records):,} records and {len(sheet_tables)} sheets to {OUTPUT}")


if __name__ == "__main__":
    main()
